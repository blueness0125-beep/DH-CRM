"""
Excel -> Supabase 임포트 스크립트
- 26년4월8일고객관리.xlsx 데이터를 Supabase DB에 삽입
- 중복 제거, 가족 그룹화, 관계 설정
"""

import openpyxl
import requests
import json
import re
from datetime import datetime, date

SUPABASE_URL = "https://ykaanrwdsgdaofzzwwwh.supabase.co"
SERVICE_ROLE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InlrYWFucndkc2dkYW9menp3d3doIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTIwOTc0NiwiZXhwIjoyMDkwNzg1NzQ2fQ.LiIbRzQ4Uk3TjZ6p6WO8zOzTR4w_SVuuLL3je57XP5g"
EXCEL_PATH = "/Users/joongkyoung/Desktop/develop/보험 고객 관리_동행지사/26년4월8일고객관리.xlsx"

HEADERS = {
    "apikey": SERVICE_ROLE_KEY,
    "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}


def supabase_insert(table: str, data: list) -> list:
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    resp = requests.post(url, headers=HEADERS, json=data)
    if resp.status_code not in (200, 201):
        raise RuntimeError(f"Insert {table} failed: {resp.status_code} {resp.text[:500]}")
    return resp.json()


def supabase_insert_batch(table: str, rows: list, batch_size: int = 100) -> list:
    results = []
    for i in range(0, len(rows), batch_size):
        batch = rows[i : i + batch_size]
        results.extend(supabase_insert(table, batch))
        print(f"  {table}: {min(i + batch_size, len(rows))}/{len(rows)} 삽입 완료")
    return results


def get_gender(ssn_back) -> str | None:
    if ssn_back is None:
        return None
    try:
        # 숫자만 추출 후 첫 자리로 판별
        digits = re.sub(r"[^\d]", "", str(ssn_back))
        if not digits:
            return None
        first_digit = int(digits[0])
    except (ValueError, IndexError):
        return None
    if first_digit in (1, 3):
        return "M"
    if first_digit in (2, 4):
        return "F"
    return None


def parse_birth_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return None


def clean_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def clean_phone(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    digits = re.sub(r"[^\d]", "", s)
    if len(digits) == 11 and digits.startswith("01"):
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10 and digits.startswith("0"):
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    # 이미 형식이 있으면 그대로
    if re.match(r"^\d{2,3}-\d{3,4}-\d{4}$", s):
        return s
    return s if s else None


def parse_car_info(car_status: str | None) -> tuple[str | None, list[str], str | None]:
    if not car_status:
        return None, [], None
    company = None
    companies = ["삼성화재", "현대해상", "메리츠", "DB손보", "KB손보", "롯데손보", "흥국화재", "한화손보"]
    for c in companies:
        if c in car_status:
            company = c
            break
    car_numbers = re.findall(r"\d+[가-힣]+\d+", car_status)
    return company, car_numbers, clean_str(car_status)


def infer_relationship(primary: dict, member: dict) -> str:
    if not primary.get("birth_date") or not member.get("birth_date"):
        return "가족"
    try:
        p_birth = datetime.strptime(primary["birth_date"], "%Y-%m-%d")
        m_birth = datetime.strptime(member["birth_date"], "%Y-%m-%d")
    except ValueError:
        return "가족"

    age_diff = (p_birth - m_birth).days / 365.25  # 양수 = member가 연상

    p_surname = primary["name"][0] if primary.get("name") else ""
    m_surname = member["name"][0] if member.get("name") else ""
    same_surname = p_surname == m_surname

    p_gender = primary.get("gender")
    m_gender = member.get("gender")

    # 배우자: 나이 차이 15세 이내
    if abs(age_diff) <= 15:
        if p_gender and m_gender and p_gender != m_gender:
            return "배우자"
        if not same_surname:
            return "배우자"

    # 자녀: member가 primary보다 15세 이상 어림
    if age_diff < -15:
        return "자녀"

    # 부모: member가 primary보다 15세 이상 연상
    if age_diff > 15:
        return "부모"

    # 형제자매: 같은 성씨, 나이 차이 15세 이내
    if same_surname and abs(age_diff) <= 15:
        return "형제/자매"

    return "가족"


def parse_row(row_vals: list, col_map: dict) -> dict:
    """컬럼 인덱스 매핑을 통해 행 파싱"""
    def get(key):
        idx = col_map.get(key)
        if idx is None:
            return None
        v = row_vals[idx] if idx < len(row_vals) else None
        return v

    ssn_back = get("ssn_back")
    phone = get("phone")
    phone2 = get("phone2")
    p = clean_phone(phone)
    p2 = clean_phone(phone2)

    return {
        "name": clean_str(get("name")),
        "birth_date": parse_birth_date(get("birth")),
        "ssn_back": clean_str(str(ssn_back)) if ssn_back is not None else None,
        "gender": get_gender(ssn_back),
        "phone": p,
        "phone_2": p2 if p and p2 and p != p2 else None,
        "home_address": clean_str(get("address")),
        "memo": clean_str(get("coverage")),
        "_car_expiry": get("car_expiry") if isinstance(get("car_expiry"), datetime) else None,
        "_car_status": clean_str(str(get("car_status"))) if get("car_status") else None,
        "_driver_info": clean_str(str(get("driver"))) if get("driver") else None,
        "_caregiver": clean_str(str(get("caregiver"))) if get("caregiver") else None,
        "_medical": clean_str(str(get("medical"))) if get("medical") else None,
        "_property": clean_str(str(get("property"))) if get("property") else None,
        "_dementia": clean_str(str(get("dementia"))) if get("dementia") else None,
        "_life": clean_str(str(get("life"))) if get("life") else None,
        "_cancer": clean_str(str(get("cancer"))) if get("cancer") else None,
    }


def dedup_key(c: dict) -> str:
    return f"{c['name']}_{c['birth_date']}"


def merge_customer(existing: dict, new: dict) -> dict:
    """기존 고객 정보에 새 정보를 병합 (비어 있는 필드만 채움)"""
    result = dict(existing)
    for k, v in new.items():
        if result.get(k) is None and v is not None:
            result[k] = v
    return result


def parse_main_sheet(ws) -> list[dict]:
    """고객관리 시트에서 가족 그룹 리스트 반환"""
    # col_map: 0-indexed
    col_map = {
        "seq": 0,        # 순
        "car_expiry": 1, # 자동차만기
        "name": 2,       # 이름
        "birth": 3,      # 주민번호(앞)
        "ssn_back": 4,   # 뒷번호
        "phone": 6,      # 전화번호
        "coverage": 7,   # 보장분석
        "car_status": 8, # 자동차보험가입현황
        "driver": 9,     # 운전자
        "caregiver": 10, # 간병인
        "address": 11,   # 주소
        "medical": 12,   # 실비
        "property": 13,  # 재물
        "dementia": 14,  # 치매
        "life": 15,      # 종신
        "cancer": 16,    # 암
        "phone2": 17,    # 전화번호2
    }

    families = []
    current_family = None
    seen_seq_codes = set()  # 중복 sequence_code 추적

    for row in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=row, column=col + 1).value for col in range(18)]
        seq_raw = row_vals[col_map["seq"]]
        seq = clean_str(str(seq_raw)) if seq_raw is not None else None

        c = parse_row(row_vals, col_map)
        if not c["name"]:
            continue

        if seq:  # 새 가족 그룹 시작
            # 중복 sequence_code는 None으로
            if seq in seen_seq_codes:
                seq = None
            else:
                seen_seq_codes.add(seq)
            current_family = {
                "sequence_code": seq,
                "primary": c,
                "members": [],
            }
            families.append(current_family)
        else:
            if current_family is not None:
                current_family["members"].append(c)
            else:
                current_family = {
                    "sequence_code": None,
                    "primary": c,
                    "members": [],
                }
                families.append(current_family)

    return families


def parse_unmanaged_sheet(ws, existing_keys: set) -> list[dict]:
    col_map = {
        "seq": 0,
        "name": 1,
        "birth": 2,
        "ssn_back": 3,
        "phone": 5,
        "address": 6,
        "medical": 7,
        "driver": 8,
        "dementia": 10,
        "life": 11,
        "cancer": 12,
    }
    customers = []
    for row in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=row, column=col + 1).value for col in range(14)]
        c = parse_row(row_vals, col_map)
        if not c["name"]:
            continue
        key = dedup_key(c)
        if key in existing_keys:
            continue
        c["_unmanaged"] = True
        customers.append(c)
    return customers


def deduplicate_families(families: list[dict]) -> list[dict]:
    """
    전체 가족 목록에서 중복 고객 제거
    - 같은 (name, birth_date)가 다른 가족에 중복으로 나타나면 첫 번째만 유지
    - 같은 가족 내 중복은 merge
    """
    global_seen: dict[str, dict] = {}  # key -> customer 정보

    deduped_families = []
    for fam in families:
        # primary 중복 체크
        p_key = dedup_key(fam["primary"])
        if p_key in global_seen:
            # primary가 중복 → 이 가족은 멤버들만 다른 가족에 흡수하거나 스킵
            # 멤버가 있으면 그들은 기존 가족에 추가 (단순화: 스킵)
            continue
        global_seen[p_key] = fam["primary"]

        # 멤버 중복 체크
        deduped_members = []
        for m in fam["members"]:
            m_key = dedup_key(m)
            if m_key in global_seen:
                continue
            global_seen[m_key] = m
            deduped_members.append(m)

        deduped_families.append({
            "sequence_code": fam["sequence_code"],
            "primary": fam["primary"],
            "members": deduped_members,
        })

    return deduped_families


def build_customer_row(c: dict, family_group_id: str | None, is_primary: bool, seq_code: str | None) -> dict:
    return {
        "sequence_code": seq_code if is_primary else None,
        "name": c["name"],
        "birth_date": c["birth_date"],
        "ssn_back": c.get("ssn_back"),
        "gender": c.get("gender"),
        "phone": c.get("phone"),
        "phone_2": c.get("phone_2"),
        "home_address": c.get("home_address"),
        "memo": c.get("memo"),
        "family_group_id": family_group_id,
        "is_primary": is_primary,
    }


def main():
    print("=== 엑셀 데이터 파싱 시작 ===")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # 1. 고객관리 시트 파싱
    ws_main = wb["고객관리"]
    families_raw = parse_main_sheet(ws_main)
    print(f"고객관리 원본: {len(families_raw)}개 그룹")

    # 2. 중복 제거
    families = deduplicate_families(families_raw)
    print(f"중복 제거 후: {len(families)}개 그룹")

    # 3. 기존 키 수집
    existing_keys = set()
    for fam in families:
        existing_keys.add(dedup_key(fam["primary"]))
        for m in fam["members"]:
            existing_keys.add(dedup_key(m))

    # 4. 미관리고객 파싱 (중복 제외)
    ws_unmanaged = wb["미관리고객"]
    unmanaged = parse_unmanaged_sheet(ws_unmanaged, existing_keys)
    print(f"미관리고객: {len(unmanaged)}명 추가")

    multi_families = [f for f in families if f["members"]]
    single_families = [f for f in families if not f["members"]]
    print(f"\n다중 구성원 가족: {len(multi_families)}개")
    print(f"단독 고객 그룹: {len(single_families)}개")

    total_main = sum(1 + len(f["members"]) for f in families)
    print(f"전체 고객 수: {total_main + len(unmanaged)}명 (고객관리 {total_main} + 미관리 {len(unmanaged)})")

    # ===== DB 삽입 =====
    print("\n=== DB 삽입 시작 ===")

    # family_groups 삽입
    family_group_rows = [{"name": f"{f['primary']['name']} 가족"} for f in multi_families]
    print(f"\nfamily_groups 삽입: {len(family_group_rows)}개")
    fg_results = supabase_insert_batch("family_groups", family_group_rows)
    for i, fam in enumerate(multi_families):
        fam["_family_group_id"] = fg_results[i]["id"]
    for fam in single_families:
        fam["_family_group_id"] = None

    # customers 삽입
    customer_rows = []
    customer_meta = []

    for fam in families:
        fg_id = fam.get("_family_group_id")
        seq = fam.get("sequence_code")
        p = fam["primary"]
        customer_rows.append(build_customer_row(p, fg_id, True, seq))
        customer_meta.append({"data": p, "family": fam})
        for m in fam["members"]:
            customer_rows.append(build_customer_row(m, fg_id, False, None))
            customer_meta.append({"data": m, "family": fam})

    for u in unmanaged:
        customer_rows.append(build_customer_row(u, None, False, None))
        customer_meta.append({"data": u, "family": None})

    print(f"\ncustomers 삽입: {len(customer_rows)}명")
    cust_results = supabase_insert_batch("customers", customer_rows)

    for i, meta in enumerate(customer_meta):
        meta["id"] = cust_results[i]["id"]
        meta["data"]["_id"] = cust_results[i]["id"]

    # customer_relationships 삽입
    rel_rows = []
    idx = 0
    for fam in families:
        primary = fam["primary"]
        idx += 1  # primary
        p_id = primary.get("_id")
        for member in fam["members"]:
            idx += 1
            m_id = member.get("_id")
            if not p_id or not m_id:
                continue
            rel_type = infer_relationship(primary, member)
            reverse_map = {
                "배우자": "배우자",
                "자녀": "부모",
                "부모": "자녀",
                "형제/자매": "형제/자매",
                "가족": "가족",
            }
            rel_rows.append({"customer_id": p_id, "related_customer_id": m_id, "relationship_type": rel_type})
            rel_rows.append({"customer_id": m_id, "related_customer_id": p_id, "relationship_type": reverse_map.get(rel_type, "가족")})

    print(f"\ncustomer_relationships 삽입: {len(rel_rows)}개")
    if rel_rows:
        supabase_insert_batch("customer_relationships", rel_rows)

    # car_insurances 삽입
    car_rows = []
    for meta in customer_meta:
        d = meta["data"]
        cust_id = meta["id"]
        car_status = d.get("_car_status")
        car_expiry = d.get("_car_expiry")
        if not car_status and not car_expiry:
            continue
        company, car_numbers, memo = parse_car_info(car_status)
        expiry_str = car_expiry.strftime("%Y-%m-%d") if isinstance(car_expiry, datetime) else None
        if car_numbers:
            for car_num in car_numbers:
                car_rows.append({"customer_id": cust_id, "car_number": car_num, "insurance_company": company, "expiry_date": expiry_str, "memo": memo})
        else:
            car_rows.append({"customer_id": cust_id, "car_number": None, "insurance_company": company, "expiry_date": expiry_str, "memo": memo})

    print(f"\ncar_insurances 삽입: {len(car_rows)}개")
    if car_rows:
        supabase_insert_batch("car_insurances", car_rows)

    # insurance_products 삽입
    ins_fields = [
        ("_medical", "실비"),
        ("_property", "재물"),
        ("_dementia", "치매"),
        ("_life", "종신"),
        ("_cancer", "암"),
        ("_driver_info", "운전자"),
        ("_caregiver", "간병인"),
    ]
    skip_vals = {"-", "N/A", "없음", "X", "x", "n/a"}
    ins_rows = []
    for meta in customer_meta:
        d = meta["data"]
        cust_id = meta["id"]
        for field, product_type in ins_fields:
            val = d.get(field)
            if val and val.strip() not in skip_vals:
                ins_rows.append({"customer_id": cust_id, "product_type": product_type, "memo": val, "status": "active"})

    print(f"\ninsurance_products 삽입: {len(ins_rows)}개")
    if ins_rows:
        supabase_insert_batch("insurance_products", ins_rows)

    print("\n=== 임포트 완료 ===")
    print(f"총 고객: {len(customer_rows)}명")
    print(f"가족 그룹: {len(multi_families)}개")
    print(f"관계: {len(rel_rows)}개")
    print(f"차량보험: {len(car_rows)}개")
    print(f"보험상품: {len(ins_rows)}개")


if __name__ == "__main__":
    main()
